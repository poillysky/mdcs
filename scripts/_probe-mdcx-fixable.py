from __future__ import annotations

import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from enrich_actor_maps import (  # noqa: E402
    KANA_RE,
    MDCX_XLSX,
    ZH_CN_PATH,
    entry_name,
    load_json,
    looks_chinese,
    needs_zh_cn,
    try_opencc_tw2s,
)

KANA = KANA_RE


def load_mdcx_full() -> dict[str, str]:
    if not MDCX_XLSX.is_file():
        return {}
    from openpyxl import load_workbook

    wb = load_workbook(MDCX_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for name in wb.sheetnames:
        cand = wb[name]
        first = next(cand.iter_rows(values_only=True))
        heads = "".join(str(c or "") for c in first[:4])
        if "日文" in heads or "中文" in heads or "原名" in heads:
            ws = cand
            break
    out: dict[str, str] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        jp = str(row[0] or "").strip()
        zh_cn = str(row[1] or "").strip()
        zh_tw = str(row[2] or "").strip()
        aliases = re.split(r"[,，、/;｜|]+", str(row[3] or ""))
        display = zh_cn or zh_tw or ""
        if not display:
            continue
        keys = {k for k in [jp, zh_cn, zh_tw, *[a.strip() for a in aliases if a.strip()]] if k}
        for key in keys:
            out.setdefault(key, display)
    wb.close()
    return out


table = load_json(ZH_CN_PATH)
missing = [k for k, v in table.items() if needs_zh_cn(k, v)]
mdcx = load_mdcx_full()

fixable = []
for k in missing:
    disp = mdcx.get(k, "")
    simp = try_opencc_tw2s(disp)
    if looks_chinese(simp) and not KANA.search(simp):
        fixable.append((k, simp))

print(f"missing={len(missing)} mdcx_fixable={len(fixable)}")
for k, v in fixable[:25]:
    print(f"  {k} -> {v}")
