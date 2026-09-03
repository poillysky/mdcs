"""补全 actors.zh-CN.json 中缺少简体中文显示名的条目。



来源（按优先级）：

1. mdcx actor_database.xlsx（仅 zh_cn / 繁转简，含别名 key）

2. 键名本身为中文 → 作显示名

3. actors.zh-TW.json 繁体 → OpenCC 转简

4. SQLite actor_profiles.mapped_name

5. 同 URL / libredmm·javdb ID 簇内择优传播中文名

6. gfriends Filetree 别名

7. 可选 --minnano N：对库内高频缺中文演员刮 minnano（慢）



用法：

  python scripts/enrich_actor_maps.py

  python scripts/enrich_actor_maps.py --dry-run

  python scripts/enrich_actor_maps.py --minnano 50

"""

from __future__ import annotations



import json

import re

import sqlite3

import sys

import time

import urllib.parse

import urllib.request

from pathlib import Path



ROOT = Path(__file__).resolve().parents[1]

MAPS_DIR = ROOT / "data" / "scrape_maps"

ZH_CN_PATH = MAPS_DIR / "actors.zh-CN.json"

ZH_TW_PATH = MAPS_DIR / "actors.zh-TW.json"

MDCX_XLSX = ROOT / "references" / "mdcx-diy" / "resources" / "userdata" / "actor_database.xlsx"

DB_PATH = ROOT / "data" / "mdcs.db"

GFRIENDS_CACHE = ROOT / "data" / "emby_actor_cache" / "gfriends_index.json"

GFRIENDS_URL = "https://raw.githubusercontent.com/gfriends/gfriends/master/Filetree.json"

MINNANO_BASE = "https://www.minnano-av.com"

MINNANO_UA = (

    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "

    "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"

)



KANA_RE = re.compile(r"[\u3040-\u30ff]")

CJK_RE = re.compile(r"[\u4e00-\u9fff]")

LATIN_RE = re.compile(r"[A-Za-z]")

ALIAS_SPLIT_RE = re.compile(r"[,，、/;｜|]+")

LIBREDMM_ID_RE = re.compile(r"/actresses/(\d+)", re.I)

JAVDB_ID_RE = re.compile(r"/actors/([a-z0-9]+)", re.I)





def atomic_write_json(path: Path, obj: object) -> None:

    path.parent.mkdir(parents=True, exist_ok=True)

    tmp = path.with_suffix(path.suffix + ".tmp")

    tmp.write_text(json.dumps(obj, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    tmp.replace(path)





def looks_chinese(name: str) -> bool:

    s = (name or "").strip()

    if not s:

        return False

    cn = len(CJK_RE.findall(s))

    kana = len(KANA_RE.findall(s))

    return cn >= 2 and cn >= kana





def score_display_name(name: str) -> int:

    """分数越高越适合作为简中显示名。"""

    s = (name or "").strip()

    if not s:

        return -10_000

    cjk = CJK_RE.findall(s)

    kana = KANA_RE.findall(s)

    score = 0

    if kana:

        score -= 120 * len(kana)

    if len(cjk) >= 2:

        score += 80 + len(cjk) * 8

    elif len(cjk) == 1:

        score += 10

    if LATIN_RE.search(s):

        score -= 40

    if looks_chinese(s) and not kana:

        score += 60

    return score





def pick_best_display(candidates: list[str]) -> str:

    best = ""

    best_score = -10_000

    seen: set[str] = set()

    for raw in candidates:

        for name in (raw, try_opencc_tw2s(raw)):

            name = name.strip()

            if not name or name in seen:

                continue

            seen.add(name)

            sc = score_display_name(name)

            if sc > best_score:

                best_score = sc

                best = name

    return best if best_score >= 50 else ""





def entry_name(entry: object) -> str:

    if isinstance(entry, str):

        return entry.strip()

    if isinstance(entry, dict):

        return str(entry.get("name") or entry.get("zh") or entry.get("title") or "").strip()

    return ""





def entry_href(entry: object) -> str:

    if not isinstance(entry, dict):

        return ""

    for k in ("javdb", "javdbUrl", "url", "link"):

        v = str(entry.get(k) or "").strip()

        if v.startswith("http"):

            return v

    return ""





def normalize_payload(name: str, href: str, prev: object) -> dict | str:

    name = name.strip()

    if not name:

        return prev if isinstance(prev, (dict, str)) else ""

    if href:

        key = "javdb" if "javdb" in href.lower() else "url"

        return {key: href, "name": name}

    if isinstance(prev, dict):

        out = dict(prev)

        out["name"] = name

        return out

    return name





def needs_zh_cn(key: str, entry: object) -> bool:

    name = entry_name(entry) or key

    if looks_chinese(name) and not KANA_RE.search(name) and name != key:

        return False

    if looks_chinese(name) and not KANA_RE.search(name) and not KANA_RE.search(key):

        return False

    return bool(KANA_RE.search(key) or KANA_RE.search(name) or not looks_chinese(name))





def load_json(path: Path) -> dict:

    if not path.is_file():

        return {}

    raw = path.read_text(encoding="utf-8").strip()

    if not raw:

        return {}

    return json.loads(raw)





def try_opencc_tw2s(text: str) -> str:

    try:

        from opencc import OpenCC  # type: ignore



        return OpenCC("t2s").convert(text)

    except Exception:

        return text





def actor_cluster_id(href: str) -> str:

    href = (href or "").strip()

    if not href:

        return ""

    m = LIBREDMM_ID_RE.search(href)

    if m:

        return f"libredmm:{m.group(1)}"

    m = JAVDB_ID_RE.search(href)

    if m:

        return f"javdb:{m.group(1).lower()}"

    return href.rstrip("/").lower()





def split_aliases(raw: str) -> list[str]:

    return [p.strip() for p in ALIAS_SPLIT_RE.split(raw or "") if p.strip()]





def load_tw2s(cn_table: dict) -> dict[str, str]:

    tw = load_json(ZH_TW_PATH)

    out: dict[str, str] = {}

    for key, entry in tw.items():

        name = entry_name(entry)

        if not looks_chinese(name) or KANA_RE.search(name):

            continue

        simp = try_opencc_tw2s(name)

        if not looks_chinese(simp) or KANA_RE.search(simp):

            continue

        if key not in cn_table or needs_zh_cn(key, cn_table.get(key, "")):

            out[key] = simp

    return out





def load_profile_names() -> dict[str, str]:

    if not DB_PATH.is_file():

        return {}

    out: dict[str, str] = {}

    conn = sqlite3.connect(DB_PATH)

    try:

        rows = conn.execute(

            "SELECT name, mapped_name FROM actor_profiles WHERE mapped_name IS NOT NULL AND mapped_name != ''"

        ).fetchall()

        for name, mapped in rows:

            name = str(name or "").strip()

            mapped = str(mapped or "").strip()

            if name and mapped and looks_chinese(mapped) and not KANA_RE.search(mapped):

                out[name] = try_opencc_tw2s(mapped)

    except sqlite3.Error:

        pass

    finally:

        conn.close()

    return out





def load_mdcx_xlsx() -> dict[str, str]:

    """xlsx 行内所有别名 key → 简中显示（不用日文列兜底）。"""

    if not MDCX_XLSX.is_file():

        return {}

    try:

        from openpyxl import load_workbook

    except ImportError:

        print("  skip mdcx xlsx: openpyxl not installed", file=sys.stderr)

        return {}



    wb = load_workbook(MDCX_XLSX, read_only=True, data_only=True)

    ws = None

    for name in wb.sheetnames:

        cand = wb[name]

        first = next(cand.iter_rows(values_only=True))

        heads = "".join(str(c or "") for c in first[:4])

        if "日文" in heads or "中文" in heads or "原名" in heads:

            ws = cand

            break

    if ws is None:

        ws = wb[wb.sheetnames[0]]



    out: dict[str, str] = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):

        if i == 0:

            continue

        jp = str(row[0] or "").strip()

        zh_cn = str(row[1] or "").strip()

        zh_tw = str(row[2] or "").strip()

        aliases = split_aliases(str(row[3] or ""))

        display = ""

        if zh_cn and looks_chinese(zh_cn) and not KANA_RE.search(zh_cn):

            display = zh_cn

        elif zh_tw:

            simp = try_opencc_tw2s(zh_tw)

            if looks_chinese(simp) and not KANA_RE.search(simp):

                display = simp

        if not display:

            continue

        keys = {k for k in [jp, zh_cn, zh_tw, *aliases] if k}

        for key in keys:

            out.setdefault(key, display)

    wb.close()

    return out





def key_name_fill(table: dict) -> dict[str, str]:

    """键名本身是中文时，纠正被写成日文的显示名。"""

    out: dict[str, str] = {}

    for key, entry in table.items():

        if not looks_chinese(key) or KANA_RE.search(key):

            continue

        display = try_opencc_tw2s(key)

        if not looks_chinese(display) or KANA_RE.search(display):

            continue

        if needs_zh_cn(key, entry):

            out[key] = display

    return out





def url_cluster_fill(table: dict) -> dict[str, str]:

    """同演员外链 ID 簇内择优传播中文名（含键名候选）。"""

    by_id: dict[str, list[tuple[str, str]]] = {}

    for key, entry in table.items():

        cid = actor_cluster_id(entry_href(entry))

        if not cid:

            continue

        name = entry_name(entry) or key

        by_id.setdefault(cid, []).append((key, name))



    updates: dict[str, str] = {}

    for _cid, items in by_id.items():

        candidates: list[str] = []

        for key, name in items:

            candidates.extend([name, key])

        best = pick_best_display(candidates)

        if not best:

            continue

        for key, _name in items:

            if needs_zh_cn(key, table.get(key, "")):

                updates[key] = best

    return updates





def gfriends_alias_map() -> dict[str, str]:

    data = None

    if GFRIENDS_CACHE.is_file():

        try:

            data = json.loads(GFRIENDS_CACHE.read_text(encoding="utf-8"))

        except json.JSONDecodeError:

            data = None

    if data is None:

        try:

            print("  downloading gfriends Filetree.json …")

            with urllib.request.urlopen(GFRIENDS_URL, timeout=120) as resp:

                data = json.loads(resp.read().decode("utf-8"))

            GFRIENDS_CACHE.parent.mkdir(parents=True, exist_ok=True)

            GFRIENDS_CACHE.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")

        except Exception as e:

            print(f"  skip gfriends: {e}", file=sys.stderr)

            return {}



    content = (data or {}).get("Content") or {}

    path_basenames: dict[str, set[str]] = {}

    for _cat, items in content.items():

        if not isinstance(items, dict):

            continue

        for filename, filepath in items.items():

            base = re.sub(r"\.(jpe?g|png|webp)$", "", filename, flags=re.I).strip()

            if not base:

                continue

            path_basenames.setdefault(filepath.lower(), set()).add(base)



    alias: dict[str, str] = {}

    for bases in path_basenames.values():

        if len(bases) < 2:

            continue

        chinese = [b for b in bases if looks_chinese(b) and not KANA_RE.search(b)]

        japanese = [b for b in bases if KANA_RE.search(b)]

        if not chinese or not japanese:

            continue

        cn = pick_best_display(chinese)

        if not cn:

            continue

        for jp in japanese:

            if needs_zh_cn(jp, jp):

                alias.setdefault(jp, cn)

    return alias





def top_missing_from_db(table: dict, limit: int) -> list[str]:

    if not DB_PATH.is_file() or limit <= 0:

        return []

    freq: dict[str, int] = {}

    conn = sqlite3.connect(DB_PATH)

    try:

        rows = conn.execute(

            "SELECT meta_json FROM scrape_cache WHERE meta_json IS NOT NULL AND meta_json != ''"

        ).fetchall()

        for (raw,) in rows:

            try:

                meta = json.loads(raw)

            except json.JSONDecodeError:

                continue

            for a in meta.get("actors") or []:

                name = str(a).strip()

                if name and needs_zh_cn(name, table.get(name, "")):

                    freq[name] = freq.get(name, 0) + 1

    finally:

        conn.close()

    return [k for k, _ in sorted(freq.items(), key=lambda kv: (-kv[1], kv[0]))[:limit]]





def fetch_url(url: str, timeout: int = 30) -> str:

    req = urllib.request.Request(url, headers={"User-Agent": MINNANO_UA})

    with urllib.request.urlopen(req, timeout=timeout) as resp:

        return resp.read().decode("utf-8", errors="replace")





def parse_minnano_aliases(html: str) -> list[str]:

    aliases: list[str] = []

    for m in re.finditer(r"別名</span>\s*</td>\s*<td[^>]*>\s*<p[^>]*>([^<]+)", html, re.I | re.S):

        block = re.sub(r"\s+", " ", m.group(1)).strip()

        for part in ALIAS_SPLIT_RE.split(block):

            part = re.sub(r"\s*（.+）$", "", part).strip()

            if part:

                aliases.append(part)

    return aliases





def minnano_lookup(name: str) -> str:

    q = urllib.parse.quote(name)

    url = f"{MINNANO_BASE}/search_result.php?search_scope=actress&search_word={q}&search=+Go+"

    try:

        html = fetch_url(url)

    except Exception:

        return ""

    if "act-profile" in html and "別名" in html:

        detail_html = html

    else:

        m = re.search(r'href="([^"]*actress\d+\.html)"', html)

        if not m:

            return ""

        detail_path = m.group(1)

        detail_url = detail_path if detail_path.startswith("http") else f"{MINNANO_BASE}/{detail_path.lstrip('/')}"

        try:

            detail_html = fetch_url(detail_url)

        except Exception:

            return ""

    candidates = parse_minnano_aliases(detail_html)

    h2 = re.search(r"<h2[^>]*>\s*(.+?)\s+（", detail_html, re.S)

    if h2:

        candidates.append(h2.group(1).strip())

    return pick_best_display(candidates)





def minnano_fill(table: dict, limit: int) -> dict[str, str]:

    names = top_missing_from_db(table, limit)

    if not names:

        return {}

    out: dict[str, str] = {}

    for i, name in enumerate(names, 1):

        print(f"  minnano [{i}/{len(names)}] {name} …")

        zh = minnano_lookup(name)

        if zh and looks_chinese(zh) and not KANA_RE.search(zh):

            out[name] = zh

        time.sleep(0.8)

    return out





def main() -> int:

    dry_run = "--dry-run" in sys.argv

    minnano_limit = 0

    for arg in sys.argv[1:]:

        if arg.startswith("--minnano"):

            parts = arg.split("=", 1)

            minnano_limit = int(parts[1] if len(parts) > 1 else "30")



    table: dict = load_json(ZH_CN_PATH)

    if not table:

        print(f"empty or missing: {ZH_CN_PATH}", file=sys.stderr)

        return 1



    pending = {k for k, v in table.items() if needs_zh_cn(k, v)}

    print(f"actors.zh-CN: {len(table)} keys, {len(pending)} need zh-CN name")



    updates: dict[str, str] = {}



    def apply_source(label: str, src: dict[str, str]) -> None:

        n = 0

        for key, name in src.items():

            if key not in table:

                continue

            if not needs_zh_cn(key, table[key]):

                continue

            name = try_opencc_tw2s(name.strip())

            if not looks_chinese(name) or KANA_RE.search(name):

                continue

            cur = updates.get(key, entry_name(table.get(key, "")))

            if score_display_name(name) <= score_display_name(cur):

                continue

            updates[key] = name

            n += 1

        print(f"  + {label}: {n}")



    print("sources:")

    apply_source("mdcx xlsx", load_mdcx_xlsx())

    apply_source("key-as-display", key_name_fill(table))

    apply_source("zh-TW→简", load_tw2s(table))

    apply_source("actor_profiles", load_profile_names())

    apply_source("url cluster", url_cluster_fill(table))

    apply_source("gfriends alias", gfriends_alias_map())

    if minnano_limit > 0:

        apply_source("minnano", minnano_fill(table, minnano_limit))



    if not updates:

        print("no updates")

        return 0



    for key, name in updates.items():

        href = entry_href(table.get(key, ""))

        table[key] = normalize_payload(name, href, table.get(key))



    still = sum(1 for k, v in table.items() if needs_zh_cn(k, v))

    print(f"updated {len(updates)} keys; still missing zh-CN: {still}")



    if dry_run:

        sample = list(updates.items())[:15]

        for k, v in sample:

            print(f"  {k} → {v}")

        if len(updates) > 15:

            print(f"  … and {len(updates) - 15} more")

        return 0



    atomic_write_json(ZH_CN_PATH, table)

    print(f"wrote {ZH_CN_PATH}")

    return 0





if __name__ == "__main__":

    raise SystemExit(main())


