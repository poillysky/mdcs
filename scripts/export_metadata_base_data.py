"""导出 MDCS 元数据基础数据：
1) 5435 ed2k → data/forum_titles.json
2) mdcx-diy xlsx → data/scrape_maps/actors|tags.*.json
"""
from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT_FORUM = ROOT / "data" / "forum_titles.json"
OUT_MAPS = ROOT / "data" / "scrape_maps"
MDCX_USERDATA = ROOT / "references" / "mdcx-diy" / "resources" / "userdata"


def atomic_write_json(path: Path, obj: object) -> None:
    """先写 .tmp 再替换，避免中断/并发导致空文件。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(obj, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    tmp.replace(path)

DSN = dict(
    host="192.168.2.38",
    port=5435,
    dbname="ed2k",
    user="postgres",
    password="postgres",
    connect_timeout=30,
)

CODE_RE = re.compile(
    r"(?<![A-Z0-9])("
    r"FC2[-_ ]?(?:PPV[-_ ]?)?\d{5,8}"
    r"|[A-Z]{2,10}[-_ ]?\d{2,5}[A-Z]?"
    r"|[A-Z]{2,10}\d{2,5}"
    r")(?![A-Z0-9])",
    re.I,
)

FILM_LABELS = (
    "影片名称代号",
    "影片名稱代號",
    "原文片名",
    "原影片名",
    "原片名称",
    "影片名称",
    "影片名稱",
    "影片名",
    "影片标题",
    "影片標題",
    "视频名称",
    "視頻名稱",
    "资源名称",
    "資源名稱",
    "作品名称",
    "作品名稱",
    "片名",
)
_FILM_RE = re.compile(
    r"[【\[]\s*(?:" + "|".join(map(re.escape, FILM_LABELS)) + r")\s*[】\]]\s*[：:]\s*([^\n\r]*)",
    re.I,
)

_DIRTY_EXACT = frozenset(
    {
        "欧美女优",
        "如题",
        "。。。",
        "…",
        "...",
        "【磁力】",
        "未知",
        "无标题",
        "untitled",
        "null",
        "none",
        "n/a",
        "na",
        "-",
        "—",
        "无",
        "暂无",
        "title",
        "filename",
        "提示信息",
        "提示",
    }
)


def normalize_code(raw: str) -> str | None:
    s = (raw or "").strip().upper().replace("_", "-").replace(" ", "-")
    if not s:
        return None
    s = re.sub(r"-+", "-", s)
    # FC2
    m = re.match(r"^(FC2)(?:-?PPV)?-?(\d{5,8})$", s)
    if m:
        return f"FC2-PPV-{m.group(2)}"
    m = re.match(r"^([A-Z]{2,10})-?(\d{2,5}[A-Z]?)$", s)
    if not m:
        return None
    prefix, num = m.group(1), m.group(2)
    # skip too generic
    if prefix in {"MP4", "MKV", "AVI", "GB", "MB", "HD", "FHD", "UHD", "BT", "ED2K"}:
        return None
    return f"{prefix}-{num}"


def extract_code(blob: str) -> str | None:
    best = None
    for m in CODE_RE.finditer(blob or ""):
        c = normalize_code(m.group(1))
        if not c:
            continue
        # prefer longer / more specific
        if best is None or len(c) > len(best):
            best = c
    return best


def extract_film_title(desc: str) -> str:
    m = _FILM_RE.search(desc or "")
    if not m:
        return ""
    t = (m.group(1) or "").strip()
    t = re.sub(r"\s+", " ", t)
    # strip trailing size markers
    t = re.sub(r"\s*[\(\[【]?[\d\.]+\s*(?:GB|MB|TB)[\)\]】]?\s*$", "", t, flags=re.I).strip()
    return t


def strip_leading_code(title: str, code: str) -> str:
    t = (title or "").strip()
    if not code:
        return t
    # 去掉开头番号及常见分隔
    pat = re.compile(
        r"^\s*" + re.escape(code).replace(r"\-", r"[\s_\-]*") + r"\s*[:：\-–—]?\s*",
        re.I,
    )
    t2 = pat.sub("", t, count=1).strip()
    return t2 or t


def is_chinese_title(title: str) -> bool:
    """判定是否为可用中文标题：有汉字、无假名、汉字占主导。"""
    t = (title or "").strip()
    if len(t) < 2:
        return False
    # 日文假名 → 非中文
    if re.search(r"[\u3040-\u309f\u30a0-\u30ff\uff66-\uff9d]", t):
        return False
    han = len(re.findall(r"[\u4e00-\u9fff]", t))
    if han < 2:
        return False
    # 拉丁字母（含西欧扩展）
    latin = len(re.findall(r"[A-Za-zÀ-ÿ]", t))
    # 汉字至少是拉丁的 1.5 倍，或拉丁很少（≤3）且汉字够多
    if latin > 3 and han < latin * 1.5:
        return False
    # 去掉空白数字标点后，汉字应占多数
    core = re.sub(r"[\s\d\W_]+", "", t, flags=re.UNICODE)
    if not core:
        return False
    han_core = len(re.findall(r"[\u4e00-\u9fff]", core))
    if han_core / len(core) < 0.55:
        return False
    return True


def is_usable_title(title: str, code: str) -> bool:
    t = strip_leading_code((title or "").strip(), code)
    if len(t) < 2 or len(t) > 180:
        return False
    if t.lower() in _DIRTY_EXACT:
        return False
    compact = re.sub(r"[\s\-_.]", "", t).upper()
    code_compact = re.sub(r"[\s\-_.]", "", code).upper()
    if compact == code_compact:
        return False
    if "http://" in t.lower() or "https://" in t.lower():
        return False
    return is_chinese_title(t)


def score_title(title: str) -> int:
    """越高越好：汉字多、无假名、拉丁少。"""
    t = title.strip()
    han = len(re.findall(r"[\u4e00-\u9fff]", t))
    latin = len(re.findall(r"[A-Za-z]", t))
    score = han * 5 + min(len(t), 40) - latin * 2
    if re.search(r"[\u3040-\u30ff]", t):
        score -= 100
    return score


def export_forum_titles() -> int:
    import psycopg2

    print("connecting 192.168.2.38:5435 …")
    conn = psycopg2.connect(**DSN)
    cur = conn.cursor()
    cur.execute(
        """
        SELECT rs.title, rs.description, er.search_string
        FROM resource_sources rs
        JOIN ed2k_resources er ON er.hash = rs.hash
        WHERE rs.description ~ '【\\s*(影片|资源|作品|原文|原片|视频)'
          AND (
            er.search_string ~* '[A-Z]{2,10}[-_ ]?[0-9]{2,5}'
            OR rs.title ~* '[A-Z]{2,10}[-_ ]?[0-9]{2,5}'
            OR rs.description ~* '[A-Z]{2,10}[-_ ]?[0-9]{2,5}'
            OR er.search_string ~* 'FC2'
            OR rs.title ~* 'FC2'
          )
        """
    )

    candidates: dict[str, list[str]] = defaultdict(list)
    n = 0
    while True:
        rows = cur.fetchmany(5000)
        if not rows:
            break
        for title, desc, search in rows:
            n += 1
            film = extract_film_title(desc or "")
            blob = " ".join([title or "", search or "", desc or ""])
            code = extract_code(blob)
            if not code or not is_usable_title(film, code):
                continue
            cleaned = strip_leading_code(film, code)
            if not is_chinese_title(cleaned):
                continue
            candidates[code].append(cleaned)
        if n % 50000 == 0:
            print(f"  scanned {n}, codes so far {len(candidates)}")

    cur.close()
    conn.close()

    out: dict[str, str] = {}
    for code, titles in candidates.items():
        # majority then score
        counts = Counter(titles)
        best = max(counts.items(), key=lambda kv: (kv[1], score_title(kv[0])))[0]
        out[code] = best

    atomic_write_json(OUT_FORUM, dict(sorted(out.items())))
    print(f"forum_titles: scanned={n}, unique={len(out)} → {OUT_FORUM}")
    return len(out)


def _split_aliases(raw: str) -> list[str]:
    parts = re.split(r"[,，、/;｜|]+", raw or "")
    return [p.strip() for p in parts if p.strip()]


def export_actor_maps() -> dict[str, int]:
    from openpyxl import load_workbook

    path = MDCX_USERDATA / "actor_database.xlsx"
    if not path.is_file():
        raise FileNotFoundError(path)
    print(f"reading {path} …")
    wb = load_workbook(path, read_only=True, data_only=True)
    # 第一张通常是演员数据库；若表头不对再扫
    ws = None
    for name in wb.sheetnames:
        cand = wb[name]
        first = next(cand.iter_rows(values_only=True))
        heads = [str(c or "") for c in first[:4]]
        joined = "".join(heads)
        if "日文" in joined or "中文" in joined or "原名" in joined:
            ws = cand
            print(f"  sheet={name} header={heads}")
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]
        print(f"  fallback sheet={wb.sheetnames[0]}")

    maps: dict[str, dict] = {"zh-CN": {}, "zh-TW": {}, "ja": {}, "en": {}}
    rows = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        rows += 1
        jp = str(row[0] or "").strip()
        zh_cn = str(row[1] or "").strip()
        zh_tw = str(row[2] or "").strip()
        aliases = _split_aliases(str(row[3] or ""))
        href = str(row[4] or "").strip() if len(row) > 4 else ""
        if jp in {"删除", "刪除"} or zh_cn in {"删除", "刪除"}:
            continue
        if not (jp or zh_cn or zh_tw):
            continue

        display_cn = zh_cn or zh_tw or jp
        display_tw = zh_tw or zh_cn or jp
        display_ja = jp or zh_cn or zh_tw
        keys = {k for k in [jp, zh_cn, zh_tw, *aliases] if k}

        def put(lang: str, name: str) -> None:
            if href.startswith("http"):
                payload: dict | str = (
                    {"name": name, "javdb": href}
                    if "javdb" in href.lower()
                    else {"name": name, "url": href}
                )
            else:
                payload = name
            for key in keys:
                maps[lang].setdefault(key, payload)

        put("zh-CN", display_cn)
        put("zh-TW", display_tw)
        put("ja", display_ja)
        put("en", display_cn)

    wb.close()
    counts = {}
    for lang, table in maps.items():
        out = OUT_MAPS / f"actors.{lang}.json"
        atomic_write_json(out, table)
        counts[f"actors.{lang}"] = len(table)
        print(f"  {out.name}: {len(table)} keys (from {rows} data rows)")
    return counts


def export_tag_maps() -> dict[str, int]:
    from openpyxl import load_workbook

    path = MDCX_USERDATA / "info_database.xlsx"
    if not path.is_file():
        raise FileNotFoundError(path)
    print(f"reading {path} …")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    maps: dict[str, dict] = {"zh-CN": {}, "zh-TW": {}, "ja": {}, "en": {}}
    rows = 0
    it = ws.iter_rows(values_only=True)
    header = [str(c or "").strip().lower() for c in next(it)]
    print("  info header:", header)
    # 期望: jp, zh_cn, zh_tw, keyword
    col = {name: i for i, name in enumerate(header)}
    i_jp = col.get("jp", 0)
    i_cn = col.get("zh_cn", 1)
    i_tw = col.get("zh_tw", 2)
    i_kw = col.get("keyword", 3)

    for row in it:
        rows += 1
        cells = [str(c or "").strip() for c in row]

        def cell(i: int) -> str:
            return cells[i] if i < len(cells) else ""

        jp, cn, tw, kw_raw = cell(i_jp), cell(i_cn), cell(i_tw), cell(i_kw)
        if jp.lower() in {"jp", "删除", "刪除"} or cn.lower() in {"zh_cn", "zh", "删除", "刪除"}:
            continue
        keywords = _split_aliases(kw_raw)
        for extra in (jp, cn, tw):
            if extra:
                keywords.append(extra)
        keywords = [k for k in dict.fromkeys(keywords) if k]
        if not keywords:
            continue
        for key in keywords:
            if cn:
                maps["zh-CN"].setdefault(key, cn)
            if tw or cn:
                maps["zh-TW"].setdefault(key, tw or cn)
            if jp or cn:
                maps["ja"].setdefault(key, jp or cn)
            if cn or jp:
                maps["en"].setdefault(key, cn or jp)

    wb.close()
    counts = {}
    for lang, table in maps.items():
        out = OUT_MAPS / f"tags.{lang}.json"
        atomic_write_json(out, table)
        counts[f"tags.{lang}"] = len(table)
        print(f"  {out.name}: {len(table)} keys (rows~{rows})")
    return counts


def main() -> int:
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("installing openpyxl…")
        import subprocess

        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])

    args = set(sys.argv[1:])
    do_all = not args or args == {"--all"}
    if do_all or "--forum" in args:
        print("=== 1) forum_titles from Postgres ===")
        export_forum_titles()
    if do_all or "--actors" in args:
        print("=== 2) actor maps from mdcx ===")
        export_actor_maps()
    if do_all or "--tags" in args:
        print("=== 3) tag maps from mdcx ===")
        export_tag_maps()
    print("done")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
