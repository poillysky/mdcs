"""Quick probe: mdcx xlsx coverage for missing actors."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from enrich_actor_maps import (  # noqa: E402
    MDCX_XLSX,
    ZH_CN_PATH,
    entry_name,
    load_json,
    load_mdcx_xlsx,
    needs_zh_cn,
)

KANA = re.compile(r"[\u3040-\u30ff]")


def load_mdcx_full() -> dict[str, str]:
  """Like export_metadata_base_data: all alias keys -> zh-CN display."""
  if not MDCX_XLSX.is_file():
    return {}
  from openpyxl import load_workbook

  wb = load_workbook(MDCX_XLSX, read_only=True, data_only=True)
  ws = None
  for name in wb.sheetnames:
    cand = wb[name]
    first = next(cand.iter_rows(values_only=True))
    heads = "".join(str(c or "") for c in first[:4])
    if "日文" in heads or "中文" in heads or "原名" in heads:
      ws = cand
      break
  if ws is None:
    ws = wb[wb.sheetnames[0]]

  out: dict[str, str] = {}
  for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
      continue
    jp = str(row[0] or "").strip()
    zh_cn = str(row[1] or "").strip()
    zh_tw = str(row[2] or "").strip()
    aliases = re.split(r"[,，、/;｜|]+", str(row[3] or ""))
    display = zh_cn or zh_tw or ""
    if not display:
      continue
    keys = {k for k in [jp, zh_cn, zh_tw, *[a.strip() for a in aliases if a.strip()]] if k}
    for key in keys:
      out.setdefault(key, display)
  wb.close()
  return out


table = load_json(ZH_CN_PATH)
missing = [k for k, v in table.items() if needs_zh_cn(k, v)]
mdcx_simple = load_mdcx_xlsx()
mdcx_full = load_mdcx_full()

hit_simple = sum(1 for k in missing if k in mdcx_simple)
hit_full = sum(1 for k in missing if k in mdcx_full)
print(f"missing={len(missing)}")
print(f"mdcx_simple hits={hit_simple} mdcx_full hits={hit_full}")

samples = ["天沢りん", "花芽ありす", "佐野りか", "橘ひなの", "Shaku Alice", "辻アリス"]
for s in samples:
  print(
    s,
    "table=", entry_name(table.get(s)),
    "mdcx=", mdcx_full.get(s, "-"),
  )
